from flask import Flask, render_template, jsonify
import pyodbc

import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file, request, redirect, url_for
import pandas as pd

app = Flask(__name__)

# Lista de filiais a considerar
FILIAIS = ["09ALFA01", "09ALFA07", "09ALFA02", "09ALFA06", "09ALFA03"]
# String para display (ou log)
FILIAIS_STR = ", ".join(FILIAIS)

TESTE_SQL = {
    "server": os.environ.get("MSSQL_HOST_TST", "192.168.0.246"),
    "database": os.environ.get("MSSQL_DB_TST", "protheus12_producao"),
    "user": os.environ.get("MSSQL_USER_TST", "sa"),
    "password": os.environ.get("MSSQL_PASSWORD_TST", "mYK#LTtiA2lu"),
    "port": int(os.environ.get("MSSQL_PORT_TST", 1433)),
}

# Configuração de Produção - Pode ser sobrescrita via ENV
PROD_SQL = {
    "server": os.environ.get("MSSQL_HOST_PROD", "192.168.0.243"),
    "database": os.environ.get("MSSQL_DB_PROD", "protheus12_producao"),
    "user": os.environ.get("MSSQL_USER_PROD", "consulta2"),
    "password": os.environ.get("MSSQL_PASSWORD_PROD", "consulta2"),
    "port": int(os.environ.get("MSSQL_PORT_PROD", 1433)),
}

def format_br(value):
    """Formata float para padrão BR (1.234,56) com 2 casas"""
    if value is None:
        return "0,00"
    # Formata como US (com virgula de milhar e ponto decimal)
    us_fmt = "{:,.2f}".format(value)
    # Inverte os caracteres
    return us_fmt.replace(",", "X").replace(".", ",").replace("X", ".")

app.jinja_env.filters['format_br'] = format_br

def pick_driver():
    drivers = pyodbc.drivers()
    for preferred in ("ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server"):
        if preferred in drivers:
            return preferred
    if not drivers:
        raise RuntimeError("Nenhum ODBC Driver para SQL Server encontrado.")
    return drivers[-1]

DRIVER = pick_driver()

def connect_sql(cfg: dict):
    conn_str = (
        f"DRIVER={{{DRIVER}}};"
        f"SERVER={cfg['server']},{cfg['port']};"
        f"DATABASE={cfg['database']};"
        f"UID={cfg['user']};"
        f"PWD={cfg['password']};"
        "TrustServerCertificate=yes;"
        "Encrypt=no;"
    )
    return pyodbc.connect(conn_str, timeout=10)

def _trim(v):
    return v.rstrip() if isinstance(v, str) else v

def get_produtos_teste():
    # Monta placeholders ?, ?
    placeholders = ",".join("?" * len(FILIAIS))
    
    sql = f"""
        SELECT B2_FILIAL, B2_COD, B2_LOCAL, B2_VATU1, B2_CM1, B2_QATU, B2_DMOV
          FROM SB2010
         WHERE B2_FILIAL IN ({placeholders})
         AND D_E_L_E_T_ = ''
         AND B2_COD <> ''
         AND (
             B2_VATU1 <> 0 
             OR B2_CM1 <> 0 
         )
         ORDER BY B2_FILIAL, B2_COD, B2_LOCAL
    """
    with connect_sql(TESTE_SQL) as conn:
        cur = conn.cursor()
        cur.execute(sql, tuple(FILIAIS))
        rows = cur.fetchall()

        # [(filial, cod, local, vatu1, cm1, qatu, dmov), ...]
        return [(_trim(r.B2_FILIAL), _trim(r.B2_COD), _trim(r.B2_LOCAL), r.B2_VATU1, r.B2_CM1, r.B2_QATU, _trim(r.B2_DMOV)) for r in rows]

def sync_to_prod(produtos):
    # Atualiza só se existir e só se mudou
    sql_update = """
        UPDATE p
           SET p.B2_VATU1 = ?,
               p.B2_CM1   = ?
          FROM SB2010 p
         WHERE p.B2_FILIAL = ?
           AND p.B2_COD    = ?
           AND ISNULL(p.D_E_L_E_T_,'') = ''
           AND (
                ISNULL(p.B2_VATU1, 0) <> ISNULL(?, 0)
             OR ISNULL(p.B2_CM1,   0) <> ISNULL(?, 0)
           )
    """

    sql_exists = """
        SELECT 1
          FROM SB2010
         WHERE B2_FILIAL = ?
           AND B2_COD    = ?
           AND ISNULL(D_E_L_E_T_,'') = ''
    """

    atualizados = 0
    nao_existem = 0

    with connect_sql(PROD_SQL) as conn:
        conn.autocommit = False
        cur = conn.cursor()

        for filial, cod, vatu1, cm1 in produtos:
            filial_key = _trim(filial)
            cod_key = _trim(cod)

            # sanity: por segurança, não deixa sincronizar outra filial sem querer
            # sanity: por segurança, não deixa sincronizar outra filial sem querer
            if filial_key not in FILIAIS:
                continue

            cur.execute(sql_exists, (filial_key, cod_key))
            if cur.fetchone() is None:
                nao_existem += 1
                continue

            # params: new_vatu1, new_cm1, filial, cod, cmp_vatu1, cmp_cm1
            cur.execute(sql_update, (vatu1, cm1, filial_key, cod_key, vatu1, cm1))
            if cur.rowcount > 0:
                atualizados += 1

        conn.commit()

    return atualizados, nao_existem, len(produtos)

from flask import Flask, render_template, jsonify, request
import math

# ... (rest of imports)

import time

# Cache Global simples
CACHE_DATA = None
CACHE_TIMESTAMP = None
# Cache para Exportação da Análise de Importação
LATEST_IMPORT_DATA = []

def get_cached_data(force_reload=False):
    global CACHE_DATA, CACHE_TIMESTAMP
    
    # Se já tem dados e não forçado, retorna cache
    if CACHE_DATA is not None and not force_reload:
        return CACHE_DATA

    print("--- [CACHE MISS] Carregando dados do SQL... ---")
    start_t = time.time()
    
    # 1. Busca dados
    test_data = get_produtos_teste()
    raw_prod = get_produtos_prod()
    # Chave agora inclui FILIAL para não misturar produtos iguais de filiais dif
    # Dict Key: (Filial, Cod, Local)
    prod_dict = {(r[0], r[1], r[2]): (r[3], r[4], r[5], r[6]) for r in raw_prod}
    
    # 2. Processa em memória
    full_data = []
    for t_filial, t_cod, t_local, t_vatu, t_cm, t_qatu, t_dmov in test_data:
        p_val = prod_dict.get((t_filial, t_cod, t_local))
        
        if p_val:
            p_vatu, p_cm, p_qatu, p_dmov = p_val
        else:
            p_vatu, p_cm, p_qatu, p_dmov = 0.0, 0.0, 0.0, ""
            
        t_vatu_f = round(float(t_vatu) if t_vatu else 0.0, 2)
        p_vatu_f = round(float(p_vatu) if p_vatu else 0.0, 2)
        t_cm_f = round(float(t_cm) if t_cm else 0.0, 2)
        p_cm_f = round(float(p_cm) if p_cm else 0.0, 2)
        
        # Quantidade (B2_QATU)
        t_qatu_f = round(float(t_qatu) if t_qatu else 0.0, 2)
        p_qatu_f = round(float(p_qatu) if p_qatu else 0.0, 2)

        # Agora a comparação pode ser exata (ou com epsilon muito baixo),
        # pois já arredondamos para o que é visível.
        diff_vatu = abs(t_vatu_f - p_vatu_f) > 0.000001
        diff_cm = abs(t_cm_f - p_cm_f) > 0.000001
        # Se quiser comparar Qtd também:
        # diff_qatu = abs(t_qatu_f - p_qatu_f) > 0.000001
        # Mas por enquanto a request foca em diff de valor? 
        # Vou assumir que QATU/DMOV é informativo, mas se diferir conta como diff?
        # User não especificou, mas geralmente conta. Vou manter só valor por enquanto para não explodir diffs
        # Se o user pedir para comparar Qtd, eu habilito.
        
        has_diff = diff_vatu or diff_cm
        
        full_data.append({
            "filial": t_filial,
            "cod": t_cod,
            "local": t_local,
            "t_vatu": t_vatu_f,
            "t_cm": t_cm_f,
            "p_vatu": p_vatu_f,
            "p_cm": p_cm_f,
            "t_qatu": t_qatu_f,
            "p_qatu": p_qatu_f,
            "t_dmov": t_dmov,
            "p_dmov": p_dmov,
            "diff_vatu": diff_vatu,
            "diff_cm": diff_cm,
            "has_diff": has_diff
        })

    print(f"--- [CACHE SET] Dados processados em {time.time() - start_t:.2f}s ---")
    CACHE_DATA = full_data
    CACHE_TIMESTAMP = time.strftime("%H:%M:%S")
    return full_data

def apply_filter(data, filter_type, filter_year, filter_filial):
    # 0. Filtro de Filial
    if filter_filial != 'all':
        target_filiais = filter_filial.split(',')
        data = [item for item in data if item['filial'] in target_filiais]

    # 1. Filtro de Ano (Year) prioritiário
    # Se filter_year != 'all', só mantemos itens onde pelo menos um dos DMOVs começa com aquele ano
    if filter_year != 'all':
        # Suporta múltiplos anos separados por vírgula (ex: "2024,2023")
        target_years = filter_year.split(',')
        
        filtered = []
        for item in data:
            # Extrai ano de Teste e Prod (ex: "20240116" -> "2024")
            y_test = item['t_dmov'][:4] if item['t_dmov'] else ""
            y_prod = item['p_dmov'][:4] if item['p_dmov'] else ""
            
            # Se UM deles coincidir com ALGUM dos anos selecionados, mantemos
            if y_test in target_years or y_prod in target_years:
                filtered.append(item)
        data = filtered

    # 2. Filtro de Diff/Emqual
    if filter_type == 'diff':
        return [item for item in data if item['has_diff']]
    elif filter_type == 'equal':
        return [item for item in data if not item['has_diff']]
    return data

@app.route("/")
def index():
    page = request.args.get('page', 1, type=int)
    filter_type = request.args.get('filter', 'all')
    filter_year = request.args.get('year', 'all')
    filter_filial = request.args.get('filial', 'all')
    force_reload = request.args.get('reload', '0') == '1'
    per_page = 100

    # Pega dados do cache (ou carrega se necessário/forçado)
    full_data = get_cached_data(force_reload=force_reload)
    
    # Extrair anos disponíveis para o select
    # Varre t_dmov e p_dmov
    years = set()
    for item in full_data:
        if item['t_dmov'] and len(item['t_dmov']) >= 4:
            years.add(item['t_dmov'][:4])
        if item['p_dmov'] and len(item['p_dmov']) >= 4:
            years.add(item['p_dmov'][:4])
    
    sorted_years = sorted(list(years), reverse=True)
    
    # Aplica filtros helper
    filtered_data = apply_filter(full_data, filter_type, filter_year, filter_filial)

    # Calculo de Totais (Baseado no filtro atual)
    totals = {
        't_vatu': sum(item['t_vatu'] for item in filtered_data),
        't_cm': sum(item['t_cm'] for item in filtered_data),
        'p_vatu': sum(item['p_vatu'] for item in filtered_data),
        'p_cm': sum(item['p_cm'] for item in filtered_data),
        't_qatu': sum(item['t_qatu'] for item in filtered_data),
        'p_qatu': sum(item['p_qatu'] for item in filtered_data),
    }

    # Paginação
    total_items = len(filtered_data)
    total_pages = math.ceil(total_items / per_page) if total_items > 0 else 1
    
    if page < 1: page = 1
    if page > total_pages: page = total_pages
    
    start = (page - 1) * per_page
    end = start + per_page
    
    paginated_data = filtered_data[start:end]

    # Prepara lista de anos selecionados para o template marcar
    selected_years = filter_year.split(',')
    selected_filiais = filter_filial.split(',')

    return render_template(
        "index.html",
        comparison_data=paginated_data,
        filiais_list=FILIAIS,
        filial_display=FILIAIS_STR, # Só p/ info header se quiser
        last_update=CACHE_TIMESTAMP,
        
        # Stats
        total_items=total_items, # Total filtrado
        total_full=len(full_data), # Total absoluto
        totals=totals, # Somas
        
        # Pagination & Filter
        page=page,
        total_pages=total_pages,
        current_filter=filter_type,
        current_year=filter_year,
        current_filial=filter_filial,
        selected_years=selected_years,
        selected_filiais=selected_filiais,
        available_years=sorted_years
    )

def get_produtos_prod():
    placeholders = ",".join("?" * len(FILIAIS))
    sql = f"""
        SELECT B2_FILIAL, B2_COD, B2_LOCAL, B2_VATU1, B2_CM1, B2_QATU, B2_DMOV
          FROM SB2010
         WHERE B2_FILIAL IN ({placeholders})
         AND D_E_L_E_T_ = ''
         AND B2_COD <> ''
         AND (
             B2_VATU1 <> 0 
             OR B2_CM1 <> 0 
         )
         ORDER BY B2_FILIAL, B2_COD, B2_LOCAL
    """
    with connect_sql(PROD_SQL) as conn:
        cur = conn.cursor()
        cur.execute(sql, tuple(FILIAIS))
        rows = cur.fetchall()
        return [(_trim(r.B2_FILIAL), _trim(r.B2_COD), _trim(r.B2_LOCAL), r.B2_VATU1, r.B2_CM1, r.B2_QATU, _trim(r.B2_DMOV)) for r in rows]

# Endpoint sync removido para este modo de comparação

from openpyxl.cell import WriteOnlyCell

import xlsxwriter

@app.route("/export_excel")
def export_excel():
    # 0. Obter dados e filtro
    filter_type = request.args.get('filter', 'all')
    filter_year = request.args.get('year', 'all')
    filter_filial = request.args.get('filial', 'all')
    data = get_cached_data() 
    if not data:
        data = []
    
    # 1. Aplicar o MEIO FILTRO que está na tela
    data = apply_filter(data, filter_type, filter_year, filter_filial)

    # 2. Configurar XlsxWriter com Constant Memory (Baixo uso de RAM)
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'constant_memory': True, 'in_memory': False})
    worksheet = workbook.add_worksheet("Comparacao SB2")

    # Estilos
    header_fmt = workbook.add_format({
        'bold': True,
        'font_color': '#FFFFFF',
        'bg_color': '#0B1220',
        'align': 'center'
    })
    
    total_fmt = workbook.add_format({
        'bold': True,
        'bg_color': '#e1e1e1',
        'num_format': '#,##0.00',
        'top': 1
    })

    diff_fmt = workbook.add_format({'bold': True, 'font_color': '#FF0000', 'num_format': '#,##0.00'})
    normal_fmt = workbook.add_format({'num_format': '#,##0.00'})
    text_fmt = workbook.add_format({})

    # 3. Cabeçalhos
    # [Filial, Prod, Local, T_VATU, T_CM, T_QATU, T_DMOV, P_VATU, P_CM, P_QATU, P_DMOV]
    headers = [
        "FILIAL", "PRODUTO", "LOCAL", 
        "T_QATU", "T_VATU", "T_CM", "T_DMOV", 
        "P_QATU", "P_VATU", "P_CM", "P_DMOV"
    ]
    for col, h in enumerate(headers):
        worksheet.write(0, col, h, header_fmt)

    # Ajuste largura colunas (aprox)
    worksheet.set_column(0, 0, 10) # Filial
    worksheet.set_column(1, 1, 20) # Produto
    worksheet.set_column(2, 2, 10) # Local
    worksheet.set_column(3, 10, 15) # Valores

    # Acumuladores de Totais
    sum_t_vatu = 0.0
    sum_t_cm = 0.0
    sum_t_qatu = 0.0
    sum_p_vatu = 0.0
    sum_p_cm = 0.0
    sum_p_qatu = 0.0

    # 4. Loop de Dados
    row_idx = 0
    for i, item in enumerate(data, start=1):
        row_idx = i
        # Conversões
        # Conversões (Agora já são floats, mas garantindo default 0.0)
        t_vatu = item.get('t_vatu', 0.0)
        t_cm = item.get('t_cm', 0.0)
        t_qatu = item.get('t_qatu', 0.0)
        t_dmov = item.get('t_dmov', "")
        
        p_vatu = item.get('p_vatu', 0.0)
        p_cm = item.get('p_cm', 0.0)
        p_qatu = item.get('p_qatu', 0.0)
        p_dmov = item.get('p_dmov', "")
        
        # Soma
        sum_t_vatu += t_vatu
        sum_t_cm += t_cm
        sum_t_qatu += t_qatu
        sum_p_vatu += p_vatu
        sum_p_cm += p_cm
        sum_p_qatu += p_qatu

        # Escrever células
        # 0=Filial, 1=Prod, 2=Local
        worksheet.write(row_idx, 0, item['filial'], text_fmt)
        worksheet.write(row_idx, 1, item['cod'], text_fmt)
        worksheet.write(row_idx, 2, item['local'], text_fmt)
        
        # TESTE [QATU, VATU, CM, DMOV]
        # Col 3..6
        worksheet.write(row_idx, 3, t_qatu, normal_fmt)
        worksheet.write(row_idx, 4, t_vatu, normal_fmt)
        worksheet.write(row_idx, 5, t_cm, normal_fmt)
        worksheet.write(row_idx, 6, t_dmov, text_fmt)
        
        # PROD [QATU, VATU, CM, DMOV]
        # Col 7..10
        # Formatação Condicional na Linha (se diff valor)
        fmt_vatu = diff_fmt if item['diff_vatu'] else normal_fmt
        fmt_cm = diff_fmt if item['diff_cm'] else normal_fmt
        
        worksheet.write(row_idx, 7, p_qatu, normal_fmt)
        worksheet.write(row_idx, 8, p_vatu, fmt_vatu)
        worksheet.write(row_idx, 9, p_cm, fmt_cm)
        worksheet.write(row_idx, 10, p_dmov, text_fmt)

    # 5. Escrever Totais na última linha
    last_row = row_idx + 1
    worksheet.write(last_row, 0, "TOTAL GERAL", total_fmt)
    worksheet.write(last_row, 1, "", total_fmt)
    worksheet.write(last_row, 2, "", total_fmt)
    
    # Test Totais
    worksheet.write(last_row, 3, sum_t_qatu, total_fmt)
    worksheet.write(last_row, 4, sum_t_vatu, total_fmt)
    worksheet.write(last_row, 5, sum_t_cm, total_fmt)
    worksheet.write(last_row, 6, "", total_fmt)

    # Prod Totais
    worksheet.write(last_row, 7, sum_p_qatu, total_fmt)
    worksheet.write(last_row, 8, sum_p_vatu, total_fmt)
    worksheet.write(last_row, 9, sum_p_cm, total_fmt)
    worksheet.write(last_row, 10, "", total_fmt)

    # 6. Fechar e Enviar
    # 6. Fechar e Enviar
    workbook.close()
    output.seek(0)
    
    filter_label = f"_{filter_type}" if filter_type != 'all' else ""
    filename = f"comparacao_sb2{filter_label}_{time.strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route("/importar")
def importar():
    return render_template("importar.html")

@app.route("/upload_analise", methods=["POST"])
def upload_analise():
    file = request.files.get('file')
    if not file:
        return "Nenhum arquivo enviado", 400
    
    # Check extension
    if not (file.filename.endswith('.xls') or file.filename.endswith('.xlsx')):
        return "Formato inválido. Use .xls ou .xlsx", 400

    try:
        # Ler Excel com Pandas
        df = pd.read_excel(file, decimal=',', thousands='.')
        
        # Limpar nomes de colunas (strip e lower para busca)
        df.columns = [str(c).strip() for c in df.columns]
        cols_lower = {c.lower(): c for c in df.columns}
        
        # 1. Detectar Coluna de CÓDIGO
        col_codigo = None
        for c in df.columns:
            if 'código' in c.lower() or 'codigo' in c.lower() or 'produto' in c.lower():
                col_codigo = c
                break
        
        # Fallback: coluna F (index 5) normal
        if not col_codigo and len(df.columns) > 5:
            col_codigo = df.columns[5]
            
        if not col_codigo:
             return "Não foi possível identificar a coluna de 'Código' ou 'Produto'.", 400

        # 2. Detectar Coluna de QUANTIDADE
        col_qty = None
        for c in df.columns:
            if 'quant' in c.lower() or 'qtd' in c.lower() or 'saldo' in c.lower():
                col_qty = c
                break
        
        # 3. Detectar Coluna de VALOR
        col_val = None
        for c in df.columns:
            # Evita "Unitário" se tiver "TOTAL" ou "VALOR"
            if 'valor' in c.lower() or 'total' in c.lower() or ('custo' in c.lower() and 'unit' not in c.lower()):
                col_val = c
                break
        # Se não achou 'Valor' ou 'Total', tenta Custo Unitário * Qtd depois? Não, melhor tentar só 'VALOR' como na imagem
        if not col_val and 'valor' in cols_lower:
            col_val = cols_lower['valor']

        # Carregar dados da Base Teste
        test_data_raw = get_produtos_teste()
        
        # Agrupar dados de Teste em Dict: Código -> {qatu: sum, vatu: sum, details: str}
        # Nota: Comparar linha a linha do excel com o TOTAL do código na base.
        test_db_map = {}
        for r in test_data_raw:
            # r = (filial, cod, local, vatu1, cm1, qatu, dmov)
            c_key = r[1].strip()
            if c_key not in test_db_map:
                test_db_map[c_key] = {'qatu': 0.0, 'vatu': 0.0, 'locais': []}
            
            test_db_map[c_key]['qatu'] += float(r[5] or 0)
            test_db_map[c_key]['vatu'] += float(r[3] or 0)
            test_db_map[c_key]['locais'].append(f"{r[0]}-{r[2]}")
            
        # Processar Lista Final
        results = []
        
        for idx, row in df.iterrows():
            code_val = str(row[col_codigo]).strip()
            
            # Dados do Excel (Importado)
            try:
                i_qatu = float(row[col_qty]) if col_qty else 0.0
            except: i_qatu = 0.0
            
            try:
                i_vatu = float(row[col_val]) if col_val else 0.0
            except: i_vatu = 0.0
            
            # Dados do Banco (Teste)
            db_entry = test_db_map.get(code_val)
            found = db_entry is not None
            
            t_qatu = db_entry['qatu'] if found else 0.0
            t_vatu = db_entry['vatu'] if found else 0.0
            details = "; ".join(db_entry['locais']) if found else ""
            
            # Diffs (com tolerância pequena)
            diff_qatu = abs(t_qatu - i_qatu) > 0.01
            diff_vatu = abs(t_vatu - i_vatu) > 0.01
            has_diff = diff_qatu or diff_vatu
            
            results.append({
                'cod': code_val,
                'desc': row.get('Descrição', row.get('Descr', '')), # Tenta pegar descrição se tiver
                'details': details,
                
                # Teste
                't_qatu': t_qatu,
                't_vatu': t_vatu,
                
                # Importado
                'i_qatu': i_qatu,
                'i_vatu': i_vatu,
                
                # Flags
                'found': found,
                'diff_qatu': diff_qatu,
                'diff_vatu': diff_vatu,
                'has_diff': has_diff,
                
                # Raw row for debugging or extra cols if needed (not sending all to keep light)
            })
            
        # Salva no cache global para exportação
        global LATEST_IMPORT_DATA
        LATEST_IMPORT_DATA = results

        return render_template(
            "importar_resultado.html",
            data=results,
            totals={
                't_qatu': sum(r['t_qatu'] for r in results),
                't_vatu': sum(r['t_vatu'] for r in results),
                'i_qatu': sum(r['i_qatu'] for r in results),
                'i_vatu': sum(r['i_vatu'] for r in results),
            }
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Erro ao processar arquivo: {str(e)}", 500

@app.route("/export_analise")
def export_analise():
    global LATEST_IMPORT_DATA
    if not LATEST_IMPORT_DATA:
        return "Nenhum dado disponível para exportação. Realize uma importação primeiro.", 400

    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})
    worksheet = workbook.add_worksheet("Resultado Analise")

    # Formatos
    header_fmt = workbook.add_format({'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#0B1220', 'align': 'center'})
    normal_fmt = workbook.add_format({'num_format': '#,##0.00'})
    diff_fmt = workbook.add_format({'bold': True, 'font_color': '#FF0000', 'num_format': '#,##0.00'})
    text_fmt = workbook.add_format({})

    # Cabeçalhos
    headers = [
        "CÓDIGO", "DESCRIÇÃO", "STATUS", 
        "QTD TESTE", "VALOR TESTE", "LOCAIS TESTE",
        "QTD IMPORTADA", "VALOR IMPORTADO"
    ]
    for col, h in enumerate(headers):
        worksheet.write(0, col, h, header_fmt)

    # Ajuste colunas
    worksheet.set_column(0, 0, 15) # Cod
    worksheet.set_column(1, 1, 35) # Desc
    worksheet.set_column(2, 2, 15) # Status
    worksheet.set_column(3, 7, 15) # Vals

    # Dados
    for i, item in enumerate(LATEST_IMPORT_DATA, start=1):
        worksheet.write(i, 0, item['cod'], text_fmt)
        worksheet.write(i, 1, item.get('desc', ''), text_fmt)
        
        status = "CADASTRO OK" if item['found'] else "NÃO EXISTE"
        worksheet.write(i, 2, status, text_fmt)
        
        # Teste
        worksheet.write(i, 3, item['t_qatu'], normal_fmt)
        worksheet.write(i, 4, item['t_vatu'], normal_fmt)
        worksheet.write(i, 5, item['details'], text_fmt)
        
        # Importado (Destaca Diff)
        fmt_qatu = diff_fmt if item['diff_qatu'] else normal_fmt
        fmt_vatu = diff_fmt if item['diff_vatu'] else normal_fmt
        
        worksheet.write(i, 6, item['i_qatu'], fmt_qatu)
        worksheet.write(i, 7, item['i_vatu'], fmt_vatu)

    workbook.close()
    output.seek(0)
    
    filename = f"resultado_analise_{time.strftime('%Y%m%d_%H%M')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=9901)
